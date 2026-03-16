#!/usr/bin/env python3
"""
Parses the S3 Volunteer Shift Schedule XLSX and generates SQL seed data.

Usage:
    python3 scripts/import_spreadsheet.py > sql/002_seed_data.sql
"""

import re
import sys
import uuid
import openpyxl
from datetime import datetime, date
from collections import Counter, defaultdict

XLSX_PATH = "/Users/jeff/Downloads/S3 Volunteer Shift Schedule.xlsx"

# ============================================================
# NAME NORMALIZATION MAP
# Maps messy spreadsheet names → canonical first_name
# ============================================================
NAME_MAP = {
    # Standard variations (lowercase → canonical)
    "alexandra": "Alexandra",
    "alexanda": "Alexandra",
    "alexandra-": "Alexandra",
    "alexandra-away": "Alexandra",
    "allison": "Allison",
    "andy": "Andy",
    "antoinette": "Antoinette",
    "beth": "Beth",
    "bethaway": "Beth",
    "bob": "Bob",
    "bonnie": "Bonnie",
    "bonie": "Bonnie",
    "carla": "Carla",
    "chioma": "Chioma",
    "clarence": "Clarence",
    "dave": "Dave",
    "david": "David",
    "dharam": "Dharam",
    "dharma": "Dharam",
    "diane": "Diane",
    "gina": "Gina",
    "hal": "Hal",
    "heidi": "Heidi",
    "hephzi": "Hephzi",
    "jake": "Jake",
    "james": "James",
    "jeff": "Jeff",
    "joe": "Joe",
    "judy": "Judy",
    "karen": "Karen",
    "laurent": "Laurent",
    "linda": "Linda",
    "lindsay": "Lindsay",
    "lorena": "Lorena",
    "loretta": "Loretta",
    "mag": "Mag",
    "mandy": "Mandy",
    "margret": "Margret",
    "marti": "Marti",
    "mart": "Marti",
    "mary": "Mary",
    "msry": "Mary",
    "matt": "Matt",
    "miranda": "Miranda",
    "nancy": "Nancy",
    "nannette": "Nannette",
    "nick": "Nick",
    "nikki": "Nikki",
    "noah": "Noah",
    "olivia": "Olivia",
    "preston": "Preston",
    "rhonda": "Rhonda",
    "robyn": "Robyn",
    "sam": "Sam",
    "saskia": "Saskia",
    "saska": "Saskia",
    "stephen": "Stephen",
    "sydney": "Sydney",
    "tangerine": "Tangerine",
    "tony": "Tony",
    "tracy": "Tracy",
    "wendy": "Wendy",
    "zayn": "Zayn",
    "zoe": "Zoe",
}

# Strings that are NOT volunteer names
IGNORE_STRINGS = {
    "thanksgiving", "thursday", "no gina", "no", "none", "closed",
    "holiday", "christmas", "new year",
}


def make_uuid(prefix_char, num):
    """Generate deterministic UUID with format 8-4-4-4-12.
    prefix_char: single hex char ('a', 'b', 'c', 'd', 'e')
    Result: a0000000-0000-0000-0000-000000000001
    """
    return f"{prefix_char}0000000-0000-0000-0000-{num:012d}"


def normalize_name(raw):
    """Extract canonical volunteer name from messy cell text."""
    if not raw or not isinstance(raw, str):
        return None, None, None

    text = raw.strip()
    if not text or len(text) < 2:
        return None, None, None

    # Remove leading punctuation
    text = re.sub(r'^[.(]+', '', text).strip()

    if text.lower() in IGNORE_STRINGS:
        return None, None, None

    # Parse status and notes from the text
    status = "attending"
    notes = None
    time_slot_hint = None

    text_lower = text.lower()

    # Check for away/out/sick
    if any(kw in text_lower for kw in ["away", "out", "absent", "sick", "not in", "recovering"]):
        status = "away"

    # Check for late
    if "late" in text_lower:
        status = "late"

    # Check for partial attendance (time ranges like "9-2", "1-6")
    time_range = re.search(r'\d{1,2}[:-]\d{1,2}(?::\d{2})?', text)
    if time_range and status == "attending":
        status = "partial"
        notes = time_range.group()

    # Check for parenthetical notes
    paren = re.search(r'\(([^)]+)\)', text)
    if paren:
        paren_text = paren.group(1).strip()
        paren_lower = paren_text.lower()
        if "am" in paren_lower:
            time_slot_hint = "morning"
        elif "aft" in paren_lower or "pm" in paren_lower or "eve" in paren_lower:
            time_slot_hint = "afternoon"
        if any(kw in paren_lower for kw in ["away", "out", "sick", "leaving", "not"]):
            status = "away"
        elif "late" in paren_lower:
            status = "late"
        elif notes is None and len(paren_text) > 2:
            notes = paren_text

    # Check for AM/morning hint
    if re.search(r'\bAM\b|\bam\b|\bmorning\b', text):
        time_slot_hint = "morning"

    # Check for PM/afternoon/evening hint
    if re.search(r'\bPM\b|\bpm\b|\bafternoon\b|\bAfternoon\b|\bAft\b|\baft\b', text):
        time_slot_hint = "afternoon"
    if re.search(r'\bEve\b|\beve\b|\bevening\b|\bEvening\b', text):
        time_slot_hint = "evening"

    # Extract the name: take the first word(s) before annotations
    # Remove common suffixes/annotations
    name_text = text
    # Remove parenthetical content
    name_text = re.sub(r'\([^)]*\)', '', name_text)
    # Remove everything after common annotation markers
    for pattern in [r'\s+away.*', r'\s+AWAY.*', r'\s+AM\b.*', r'\s+am\b.*',
                    r'\s+PM\b.*', r'\s+pm\b.*', r'\s+out\b.*', r'\s+late\b.*',
                    r'\s+afternoon.*', r'\s+Afternoon.*', r'\s+Aft\b.*',
                    r'\s+morning.*', r'\s+Morning.*',
                    r'\s+Eve\b.*', r'\s+eve\b.*', r'\s+evening.*',
                    r'\s+absent.*', r'\s+sick.*', r'\s+not\b.*',
                    r'\s+recovering.*', r'\s+still.*', r'\s+in\s+\w+',
                    r'\s+arrive.*', r'\s+working.*', r'\s+TBD.*',
                    r'\s+\d{1,2}[-:]\d.*',  # Time ranges
                    r'\s*-\s*(away|eve|evening|aft|afternoon|am|morning|out|sick).*',
                    ]:
        name_text = re.sub(pattern, '', name_text, flags=re.IGNORECASE)

    # Clean up
    name_text = name_text.strip().rstrip('-').rstrip('.').strip()

    if not name_text or len(name_text) < 2:
        return None, None, None

    # Normalize via map
    canonical = NAME_MAP.get(name_text.lower())
    if canonical is None:
        # Try first word only
        first_word = name_text.split()[0].rstrip('-').rstrip('.')
        canonical = NAME_MAP.get(first_word.lower())

    if canonical is None:
        # Unknown name - skip with warning
        print(f"-- WARNING: Unknown name '{name_text}' from '{raw}'", file=sys.stderr)
        return None, None, None

    return canonical, status, (notes, time_slot_hint)


def parse_spreadsheet():
    """Parse the XLSX and return structured data."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Schedule updates"]

    # Collect all entries
    entries = []  # (date, day_of_week, column, name, status, notes, time_slot_hint)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        day_cell = row[0].value  # Column A: day name
        date_cell = row[1].value  # Column B: date

        if not day_cell or not date_cell:
            continue

        if isinstance(date_cell, datetime):
            shift_date = date_cell.date()
        elif isinstance(date_cell, date):
            shift_date = date_cell
        else:
            continue

        # Skip weekends
        dow = shift_date.weekday()  # 0=Monday, 4=Friday
        if dow > 4:
            continue

        # Process volunteer cells (columns C onwards, index 2+)
        for c in row[2:]:
            if c.value is None:
                continue

            canonical, status, extra = normalize_name(str(c.value))
            if canonical is None:
                continue

            notes_text = extra[0] if extra else None
            time_slot_hint = extra[1] if extra else None

            entries.append({
                "date": shift_date,
                "day_of_week": dow,
                "column": c.column,
                "name": canonical,
                "status": status,
                "notes": notes_text,
                "time_slot_hint": time_slot_hint,
            })

    return entries


def infer_assignments(entries):
    """
    Infer regular shift assignments from frequency analysis.
    Returns {(name, day_of_week): time_slot}
    """
    # Count how often each volunteer appears on each day
    day_counts = defaultdict(Counter)  # name → {day_of_week: count}
    slot_hints = defaultdict(Counter)  # (name, day) → {time_slot: count}

    for e in entries:
        day_counts[e["name"]][e["day_of_week"]] += 1
        if e["time_slot_hint"]:
            slot_hints[(e["name"], e["day_of_week"])][e["time_slot_hint"]] += 1

    assignments = {}
    for name, days in day_counts.items():
        for dow, count in days.items():
            # Only count as regular if they appear at least 3 times on this day
            if count < 3:
                continue

            # Determine time slot from hints
            hints = slot_hints.get((name, dow), Counter())
            if hints:
                slot = hints.most_common(1)[0][0]
            else:
                slot = "afternoon"  # Default: most common shift

            assignments[(name, dow)] = slot

    return assignments


def determine_active(entries, cutoff_date=date(2026, 1, 1)):
    """Determine which volunteers are active based on last seen date."""
    last_seen = defaultdict(lambda: date.min)
    for e in entries:
        if e["date"] > last_seen[e["name"]]:
            last_seen[e["name"]] = e["date"]

    active = {}
    for name, last in last_seen.items():
        active[name] = last >= cutoff_date

    return active, last_seen


def sql_str(s):
    """Escape string for SQL."""
    if s is None:
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def generate_sql(entries):
    """Generate the complete seed SQL."""
    # Get all unique names
    all_names = sorted(set(e["name"] for e in entries))
    active_map, last_seen = determine_active(entries)
    assignments = infer_assignments(entries)

    # Assign deterministic UUIDs
    vol_uuids = {}
    for i, name in enumerate(all_names, 1):
        vol_uuids[name] = make_uuid("a", i)

    # Shift UUIDs (15 shifts: 5 days × 3 slots)
    shift_uuids = {}
    shift_num = 0
    days = ["monday", "tuesday", "wednesday", "thursday", "friday"]
    slots = ["morning", "afternoon", "evening"]
    for dow in range(5):
        for slot in slots:
            shift_num += 1
            shift_uuids[(dow, slot)] = make_uuid("b", shift_num)

    lines = []
    lines.append("-- Zencare Seed Data")
    lines.append("-- Generated by scripts/import_spreadsheet.py")
    lines.append(f"-- Source: S3 Volunteer Shift Schedule.xlsx")
    lines.append(f"-- Generated: {datetime.now().isoformat()}")
    lines.append(f"-- Volunteers: {len(all_names)}, Active: {sum(1 for v in active_map.values() if v)}")
    lines.append("")

    # ---- Volunteers ----
    lines.append("-- ============================================================")
    lines.append("-- VOLUNTEERS")
    lines.append("-- ============================================================")
    lines.append("")

    for name in all_names:
        uid = vol_uuids[name]
        is_active = "true" if active_map.get(name, False) else "false"
        is_admin = "true" if name == "Joe" else "false"
        lines.append(
            f"INSERT INTO volunteers (id, first_name, is_active, is_admin) VALUES "
            f"('{uid}'::uuid, {sql_str(name)}, {is_active}, {is_admin});"
        )

    lines.append("")

    # ---- Shifts ----
    lines.append("-- ============================================================")
    lines.append("-- SHIFTS (Mon-Fri x Morning/Afternoon/Evening)")
    lines.append("-- ============================================================")
    lines.append("")

    for dow in range(5):
        for slot in slots:
            uid = shift_uuids[(dow, slot)]
            lines.append(
                f"INSERT INTO shifts (id, day_of_week, time_slot) VALUES "
                f"('{uid}'::uuid, {dow}, '{slot}');"
            )

    lines.append("")

    # ---- Shift Assignments ----
    lines.append("-- ============================================================")
    lines.append("-- SHIFT ASSIGNMENTS (inferred from frequency)")
    lines.append("-- ============================================================")
    lines.append("")

    assign_num = 0
    for (name, dow), slot in sorted(assignments.items(), key=lambda x: (x[0][0], x[0][1])):
        assign_num += 1
        uid = make_uuid("c", assign_num)
        vol_id = vol_uuids[name]
        shift_id = shift_uuids[(dow, slot)]
        is_active = "true" if active_map.get(name, False) else "false"
        lines.append(
            f"INSERT INTO shift_assignments (id, shift_id, volunteer_id, is_active) VALUES "
            f"('{uid}'::uuid, '{shift_id}'::uuid, '{vol_id}'::uuid, {is_active});"
        )

    lines.append("")

    # ---- Preferred Shifts (same as assignments for active volunteers) ----
    lines.append("-- ============================================================")
    lines.append("-- PREFERRED SHIFTS (matching assignments for active volunteers)")
    lines.append("-- ============================================================")
    lines.append("")

    pref_num = 0
    for (name, dow), slot in sorted(assignments.items(), key=lambda x: (x[0][0], x[0][1])):
        if not active_map.get(name, False):
            continue
        pref_num += 1
        uid = make_uuid("d", pref_num)
        vol_id = vol_uuids[name]
        lines.append(
            f"INSERT INTO preferred_shifts (id, volunteer_id, day_of_week, time_slot) VALUES "
            f"('{uid}'::uuid, '{vol_id}'::uuid, {dow}, '{slot}');"
        )

    lines.append("")

    # ---- Attendance (historical records for non-default status) ----
    lines.append("-- ============================================================")
    lines.append("-- ATTENDANCE (historical - only non-default statuses)")
    lines.append("-- ============================================================")
    lines.append("")

    att_num = 0
    # Only generate attendance for entries with non-attending status
    for e in sorted(entries, key=lambda x: (x["date"], x["name"])):
        if e["status"] == "attending":
            continue

        # Find the shift for this entry
        slot = e.get("time_slot_hint")
        if not slot:
            # Look up their assignment for this day
            assigned_slot = assignments.get((e["name"], e["day_of_week"]))
            slot = assigned_slot or "afternoon"

        shift_id = shift_uuids.get((e["day_of_week"], slot))
        if not shift_id:
            continue

        att_num += 1
        uid = make_uuid("e", att_num)
        vol_id = vol_uuids[e["name"]]
        date_str = e["date"].isoformat()
        notes_sql = sql_str(e["notes"])

        lines.append(
            f"INSERT INTO attendance (id, shift_id, volunteer_id, shift_date, status, notes) VALUES "
            f"('{uid}'::uuid, '{shift_id}'::uuid, '{vol_id}'::uuid, '{date_str}', "
            f"'{e['status']}', {notes_sql}) ON CONFLICT DO NOTHING;"
        )

    lines.append("")
    lines.append(f"-- Total: {len(all_names)} volunteers, 15 shifts, "
                 f"{assign_num} assignments, {pref_num} preferences, {att_num} attendance records")

    return "\n".join(lines)


if __name__ == "__main__":
    print("-- Parsing spreadsheet...", file=sys.stderr)
    entries = parse_spreadsheet()
    print(f"-- Found {len(entries)} entries", file=sys.stderr)
    sql = generate_sql(entries)
    print(sql)
    print("-- Done!", file=sys.stderr)
